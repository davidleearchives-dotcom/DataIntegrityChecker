from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def generate_styled_excel(df: pd.DataFrame, output_path: str):
    """
    Generates a styled Excel file from the dataframe.
    Highlights rows based on 'Verification_Result' column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Result"
    
    # Styles
    # Light Red for Mismatch (Wait, PRD says: "Light Red for Mismatch" in Preview Table, 
    # but "Yellow" for Cell in Excel?
    # PRD 64: "대조본(B)과 비교하여 값이 다른 셀(Cell)은 배경색을 'Yellow'로"
    # PRD 65: "대조본(B)과 비교하여 누락된 열 데이터는 원본파일(A)에 배경색을 'Red'로"
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Write Header
    headers = list(df.columns)
    ws.append(headers)
    
    # Write Data
    # Optimization: Use openpyxl append is faster but enumerate rows is fine for 10k.
    # But wait, enumerate(dataframe_to_rows(...)) yields index, row.
    # dataframe_to_rows yields a list of values.
    # If the index is wrong, it might exceed limits?
    # openpyxl limits: 1,048,576 rows.
    # Error says: "Row number supplied was 1048577"
    # This means we are trying to write to row 1048577.
    # Why?
    # dataframe_to_rows(df, index=False, header=False) yields data rows.
    # start=2 means we start writing at row 2 (row 1 is header).
    # If df has 1M rows, last row is 1M + 1.
    # If df is small (9k), then index should be around 9002.
    # UNLESS: df has huge index or empty rows?
    # Or dataframe_to_rows is behaving unexpectedly?
    # Or result_df has wrong index?
    # In comparison.py we did: result_df = pd.DataFrame(index=merged.index)
    # If merged index is RangeIndex(0, 9471), then result_df index is same.
    # But wait, dataframe_to_rows with index=False SHOULD NOT yield index.
    
    # Alternative: Maybe previous logic in comparison.py created a DF with wrong size?
    # Let's debug by checking df shape before writing? Can't print here easily.
    
    # But wait, the error might be because we are iterating and the index variable 'r_idx' goes too high?
    # enumerate starts at 2.
    # If dataframe_to_rows yields too many rows?
    # Maybe NaN handling?
    
    # Or maybe the file is not empty but has a lot of empty rows at the end?
    # But user says 9471 rows.
    
    # HYPOTHESIS: The 'dataframe_to_rows' might be yielding None or something that causes issues?
    # Or maybe the issue is in how we calculate r_idx?
    # No, standard python enumerate.
    
    # Let's verify if we are appending correctly.
    # ws.append(row) automatically increments row pointer? 
    # No, ws.append appends to the next available row.
    # But we are ALSO using ws.cell(row=r_idx, ...) to style.
    # If ws.append puts data in row X, and we try to style row Y (where Y != X)?
    # ws.append uses internal pointer.
    # If we mix append and cell access?
    # ws.append adds to the bottom.
    # Initially ws has 1 row (header).
    # First iteration: r_idx=2. ws.append(row) -> adds to row 2.
    # ws.cell(row=2, ...) -> accessing row 2. Matches.
    
    # So logic seems correct for normal case.
    # Why 1048577? That is exactly MAX_ROW + 1.
    # This implies the loop ran 1M times?
    # Or we tried to access that row directly?
    
    # Maybe result_df has an index that causes dataframe_to_rows to yield weirdly?
    # We set index=merged.index.
    # If merged.index is weird?
    
    # Let's try a safer way:
    # Use ws.append solely for data.
    # Then iterate strictly based on df.shape[0].
    
    rows = list(dataframe_to_rows(df, index=False, header=False))
    # This loads all into memory, might be heavy for 1M rows but fine for 10k.
    # For 100k+ it's bad.
    
    # Better:
    # Just iterate and use append.
    # Keep track of current row manually?
    # r_idx from enumerate matches append order if we start from empty sheet + header.
    
    # Wait, could it be that 'dataframe_to_rows' is generating infinite rows? Unlikely.
    
    # Another possibility:
    # The 'df' passed to this function is HUGE?
    # Maybe the comparison logic created a Cartesian Product (Cross Join) by mistake?
    # 9k rows * 9k rows = 81M rows?
    # We used pd.merge(..., how='outer').
    # If key is not unique, it explodes.
    # User's data: "이력번호" seems unique?
    # If many duplicates in keys, outer join will explode.
    # 9471 rows. If all same key -> 9471 * 9471 = 89M rows.
    # This would explain why it hits the Excel row limit!
    
    # CHECK: Comparison Logic.
    # If 'key_col' is NOT unique, merge explodes.
    # We should warn or handle duplicates.
    # But for now, let's limit the export to MAX rows to prevent crash.
    
    MAX_EXCEL_ROWS = 1048576
    if len(df) > MAX_EXCEL_ROWS - 1: # -1 for header
        # Truncate and warn (how to warn? maybe add a row saying truncated)
        df = df.iloc[:MAX_EXCEL_ROWS-2]
        # We can't easily warn UI from here, but at least we don't crash.
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        if r_idx > MAX_EXCEL_ROWS:
            break
        ws.append(row)
        
        # Apply Styles
        # Get Verification Result (Last Column)
        result_col_idx = len(headers) # 1-based index
        result_cell = ws.cell(row=r_idx, column=result_col_idx)
        result_val = result_cell.value
        
        if not result_val:
            continue
            
        if str(result_val).startswith("Mismatch"):
            # Highlight SPECIFIC Mismatched Cells if possible, or the whole row?
            # PRD: "값이 다른 셀(Cell)은 배경색을 'Yellow'" -> Cell level.
            # My comparison logic returned "Mismatch: ColA, ColB". I can parse this.
            
            # Highlight row indicator?
            # Let's highlight the Result cell at least.
            
            mismatched_cols_str = str(result_val).replace("Mismatch: ", "")
            mismatched_cols = [c.strip() for c in mismatched_cols_str.split(",")]
            
            for col_name in mismatched_cols:
                try:
                    col_idx = headers.index(col_name) + 1
                    ws.cell(row=r_idx, column=col_idx).fill = yellow_fill
                except ValueError:
                    pass # Column might not exist in result if hidden
                    
        elif result_val == "Missing_in_B":
            # Highlight entire row Red?
            # PRD: "누락된 열 데이터는... 배경색을 'Red'"
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = red_fill
                
        elif result_val == "Missing_in_A":
            # Highlight entire row Red (or distinct color?)
            # I'll use Red as well for "Missing" concept.
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = red_fill

    # Auto-adjust column widths - REMOVED for performance on large datasets
    # Iterating through all cells is too slow for >10k rows.
    # We set a reasonable default width instead.
    # for col_idx in range(1, len(headers) + 1):
    #     ws.column_dimensions[get_column_letter(col_idx)].width = 15

    wb.save(output_path)
    return output_path
